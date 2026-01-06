import os
import re
import json
import time
import glob
import argparse
import threading
import requests
from datetime import datetime

import psutil
from openpyxl import Workbook, load_workbook

try:
    import pynvml
    NVML_OK = True
except Exception:
    NVML_OK = False


BASE_URL = "http://localhost:8000"

SHEET_RESULTS = "results"
SHEET_MONITOR = "monitor"
SHEET_SUMMARY = "summary"

MODEL_CONFIGS = [
    {
        "model": "gemini-2.5-flash",
        "api_key": "AIzaSyCIfhq-bc_FQMroNIBb3ZVo-eTFDWNIabE",
        "top_k": 50,
        "temperature": None,
    },
]

REQUEST_TIMEOUT = 180
SLEEP_BETWEEN_CALLS = 2
SAVE_EVERY_N_ROWS = 1
MONITOR_INTERVAL_SEC = 1.0


class ResourceMonitor:
    def __init__(self, interval=1.0, pid=None):
        self.interval = interval
        self.pid = pid or os.getpid()
        self.samples = []
        self.stop_event = threading.Event()
        self.proc = psutil.Process(self.pid)

        self.gpu_handle = None
        if NVML_OK:
            try:
                pynvml.nvmlInit()
                self.gpu_handle = pynvml.nvmlDeviceGetHandleByIndex(0)
            except Exception:
                self.gpu_handle = None

    def start(self):
        self.thread = threading.Thread(target=self.run, daemon=True)
        self.thread.start()

    def stop(self):
        self.stop_event.set()
        self.thread.join(timeout=5)
        if NVML_OK:
            try:
                pynvml.nvmlShutdown()
            except Exception:
                pass

    def gpu_stats(self):
        if not self.gpu_handle:
            return dict.fromkeys(
                ["gpu_util", "gpu_mem_used", "gpu_mem_total", "gpu_temp", "gpu_power"], ""
            )
        try:
            util = pynvml.nvmlDeviceGetUtilizationRates(self.gpu_handle)
            mem = pynvml.nvmlDeviceGetMemoryInfo(self.gpu_handle)
            temp = pynvml.nvmlDeviceGetTemperature(
                self.gpu_handle, pynvml.NVML_TEMPERATURE_GPU
            )
            try:
                power = pynvml.nvmlDeviceGetPowerUsage(self.gpu_handle) / 1000
            except Exception:
                power = ""
            return {
                "gpu_util": util.gpu,
                "gpu_mem_used": round(mem.used / 1024**2, 2),
                "gpu_mem_total": round(mem.total / 1024**2, 2),
                "gpu_temp": temp,
                "gpu_power": round(power, 2) if power != "" else "",
            }
        except Exception:
            return dict.fromkeys(
                ["gpu_util", "gpu_mem_used", "gpu_mem_total", "gpu_temp", "gpu_power"], ""
            )

    def run(self):
        psutil.cpu_percent(interval=None)
        self.proc.cpu_percent(interval=None)

        while not self.stop_event.is_set():
            mem = psutil.virtual_memory()
            row = {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "cpu_total": psutil.cpu_percent(interval=None),
                "ram_used": round(mem.used / 1024**2, 2),
                "ram_total": round(mem.total / 1024**2, 2),
                "proc_cpu": self.proc.cpu_percent(interval=None),
                "proc_rss": round(self.proc.memory_info().rss / 1024**2, 2),
            }
            row.update(self.gpu_stats())
            self.samples.append(row)
            time.sleep(self.interval)


def extract_sequence(text: str) -> str:
    m = re.search(r"```[\s\S]*?\n([\s\S]*?)```", text or "")
    seq = m.group(1) if m else (text or "")
    seq = seq.strip()
    if seq.startswith(">"):
        seq = "\n".join(l for l in seq.splitlines() if not l.startswith(">"))
    return seq.replace("\n", "").replace(" ", "")


def ensure_workbook(path: str, sheet_name: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        if ws.max_row == 1 and all(c.value is None for c in ws[1]):
            pass
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append([
            "timestamp", "model", "json_file", "item_index", "task",
            "protein_accession", "question", "seq_len",
            "top_k", "temperature",
            "protein_ids", "answer", "expected_output",
            "status", "error"
        ])
        wb.save(path)
    return wb, ws


def rag_call(model, api_key, question, sequence, top_k, temperature):
    payload = {
        "model": model,
        "api_key": api_key,
        "question": question,
        "sequence": sequence,
        "top_k": top_k,
    }
    if temperature is not None:
        payload["temperature"] = temperature

    r = requests.post(f"{BASE_URL}/rag_order", json=payload, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    j = r.json()
    return j.get("answer", ""), j.get("protein_ids", [])


def iter_items_seq(input_dir: str):
    # many json files under a dir; each file is a list of items: instruction, input (sequence in code block), output, metadata{task, protein_accession}
    json_files = glob.glob(os.path.join(input_dir, "**/*.json"), recursive=True)
    for jf in json_files:
        with open(jf, "r", encoding="utf-8") as f:
            items = json.load(f)

        for idx, item in enumerate(items, start=1):
            meta = item.get("metadata", {}) or {}
            question = (item.get("instruction") or "").strip()
            sequence = extract_sequence(item.get("input", ""))
            expected = item.get("output", "")
            yield {
                "json_file": os.path.relpath(jf, input_dir),
                "item_index": idx,
                "task": meta.get("task", ""),
                "protein_accession": meta.get("protein_accession", ""),
                "question": question,
                "sequence": sequence,
                "expected": expected,
            }


def iter_items_noseq(input_json: str):
    # one json file with a list of items: question, answer, source sequence must be "" always.
    with open(input_json, "r", encoding="utf-8") as f:
        items = json.load(f)

    base = os.path.basename(input_json)
    for idx, item in enumerate(items, start=1):
        question = (item.get("question") or "").strip()
        expected = item.get("answer", "")
        protein_accession = item.get("source", "")
        yield {
            "json_file": base,
            "item_index": idx,
            "task": "RAG result",
            "protein_accession": protein_accession,
            "question": question,
            "sequence": "",
            "expected": expected,
        }


def write_monitor_and_summary(wb, monitor: ResourceMonitor):
    # MONITOR SHEET
    if monitor.samples:
        if SHEET_MONITOR in wb.sheetnames:
            del wb[SHEET_MONITOR]
        ws_m = wb.create_sheet(SHEET_MONITOR)
        ws_m.append(list(monitor.samples[0].keys()))
        for s in monitor.samples:
            ws_m.append(list(s.values()))

        # SUMMARY SHEET
        if SHEET_SUMMARY in wb.sheetnames:
            del wb[SHEET_SUMMARY]
        ws_s = wb.create_sheet(SHEET_SUMMARY)
        ws_s.append(["metric", "value"])

        def avg(k):
            v = [x.get(k) for x in monitor.samples if isinstance(x.get(k), (int, float))]
            return round(sum(v) / len(v), 3) if v else ""

        def mx(k):
            v = [x.get(k) for x in monitor.samples if isinstance(x.get(k), (int, float))]
            return max(v) if v else ""

        metrics = [
            ("avg_cpu_total", avg("cpu_total")),
            ("max_cpu_total", mx("cpu_total")),
            ("avg_proc_cpu", avg("proc_cpu")),
            ("max_proc_cpu", mx("proc_cpu")),
            ("avg_gpu_util", avg("gpu_util")),
            ("max_gpu_util", mx("gpu_util")),
            ("max_gpu_mem_used", mx("gpu_mem_used")),
            ("max_gpu_temp", mx("gpu_temp")),
        ]
        for m in metrics:
            ws_s.append(list(m))


def run_all(mode: str, input_dir: str | None, input_json: str | None, output_xlsx: str):
    wb, ws = ensure_workbook(output_xlsx, SHEET_RESULTS)

    monitor = ResourceMonitor(interval=MONITOR_INTERVAL_SEC)
    monitor.start()

    written = 0

    try:
        if mode == "seq":
            if not input_dir:
                raise ValueError("--input-dir is required for mode=seq")
            item_iter = iter_items_seq(input_dir)
        elif mode == "noseq":
            if not input_json:
                raise ValueError("--input-json is required for mode=noseq")
            item_iter = iter_items_noseq(input_json)
        else:
            raise ValueError("mode must be one of: seq, noseq")

        for cfg in MODEL_CONFIGS:
            for it in item_iter:
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                status, err = "ok", ""
                answer, protein_ids = "", ""

                try:
                    answer, pids = rag_call(
                        cfg["model"],
                        cfg["api_key"],
                        it["question"],
                        it["sequence"],
                        cfg["top_k"],
                        cfg["temperature"],
                    )
                    protein_ids = ",".join(pids)
                except Exception as e:
                    status = "error"
                    err = str(e)

                ws.append([
                    ts,
                    cfg["model"],
                    it["json_file"],
                    it["item_index"],
                    it["task"],
                    it["protein_accession"],
                    it["question"],
                    len(it["sequence"]),
                    cfg["top_k"],
                    cfg["temperature"] or "",
                    protein_ids,
                    answer,
                    it["expected"],
                    status,
                    err,
                ])

                written += 1
                if written % SAVE_EVERY_N_ROWS == 0:
                    wb.save(output_xlsx)

                time.sleep(SLEEP_BETWEEN_CALLS)

    finally:
        monitor.stop()
        write_monitor_and_summary(wb, monitor)
        wb.save(output_xlsx)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mode", choices=["seq", "noseq"], required=True)
    p.add_argument("--input-dir", default=None, help="Required for --mode seq")
    p.add_argument("--input-json", default=None, help="Required for --mode noseq")
    p.add_argument("--output-xlsx", default=None, help="Optional output path")
    args = p.parse_args()

    if args.output_xlsx:
        out = args.output_xlsx
    else:
        if args.mode == "seq":
            if not args.input_dir:
                raise ValueError("--input-dir is required for mode=seq")
            out = os.path.join(args.input_dir, "RAG_Sequence_Search_Analysis.xlsx")
        else:
            if not args.input_json:
                raise ValueError("--input-json is required for mode=noseq")
            out = os.path.join(os.path.dirname(args.input_json), "RAG_Analysis.xlsx")

    run_all(args.mode, args.input_dir, args.input_json, out)


if __name__ == "__main__":
    main()
